
import time
start= time.time()
# import logon
# session=logon.Autosap()
import datetime,pytz
 
# tz=pytz.timezone('Asia/Kolkata')
tz=pytz.timezone('Europe/London')
import openpyxl
import shutil
from openpyxl.styles import PatternFill

import os,sys
import pandas as pd
import linktester
# df= pd.read_csv(r"C:\Users\nc92j2918\Documents\OneDrive - \autosap\todo1.csv")
# print(df.iloc[0]["name"])
# print(df)

# copying xl files
stamp = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d-%H-%M-%S")
# stamp = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
path=os.getcwd()+f"\excel data\\autosap {stamp}\\"
os.mkdir(path)
# sys.stdout = open(path+"LOG"+stamp+".txt", "w")

original = os.getcwd() +'\og3_w4.xlsx'
target =path+'Autosaplinkexcel_sheet-'+str(stamp)+'.xlsx' 
shutil.copyfile(original, target)
#ssprint(target)
xfile = openpyxl.load_workbook(target)
# global sheet
# sheet = xfile.get_sheet_by_name('sheet1')
sheet = xfile['Sheet1']
redFill = PatternFill(patternType='solid',fgColor='FF4040') #red
greenFill = PatternFill(patternType='solid',fgColor='228B22') #green
df= pd.read_excel(target,header=6,engine='openpyxl')
# print(df)
a=  ['SNO','SYSTEM','URL','STATUS','REMARKS']
# a=[]
# print(a)
b=0
xfile = openpyxl.load_workbook(target)
sheet = xfile['Sheet1']
for i in df.index:
    # print(df[a[1]][i],df[a[4]][i])
    a=linktester.linktester(df['URL'][i])
    print(a,i)
    if a :
        sheet.cell(row=i+8,column=4,value="PASS").fill=greenFill
        # sheet[f"D{i+8}"] = "PASS"
        # sheet[f"D{i+8}"].fill = PatternFill(patternType='solid',fgColor='228B22')
    else :
        b+=1
        sheet.cell(row=i+8,column=4,value="FAIL").fill=redFill
        # sheet[f"D{i+8}"] = "FAIL"
        # sheet[f"D{i+8}"].fill = PatternFill(patternType='solid',fgColor='FF4040')
    
from datetime import timedelta
tm=datetime.datetime.now(tz=tz) + timedelta(hours=1)
# sheet["B3"] = tm.strftime("%d.%m.%y")
# sheet["B6"] = datetime.datetime.now(tz=tz).strftime("%H:%M:%S")
# sheet["B4"] = tm.strftime("%H:%M:%S")
# sheet["B5"] = datetime.datetime.now(tz=tz).strftime("%d.%m.%y")
# sheet["E6"] = status
print("Filling Headers")
sheet.cell(row=3,column=2).value = stamp[:10]
sheet.cell(row=4,column=2).value = f'{stamp[11:]} GMT'
sheet.cell(row=5,column=2).value = tm.strftime("%d.%m.%y")
sheet.cell(row=6,column=2).value = tm.strftime("%H:%M:%S")
status = "Green" if b==0 else "Red"

end= time.time()
run_time=round(end-start)
print("Time Taken to run :",run_time)
# sheet["E4"] = f"{a} sec."
# sheet["E5"] = stamp
# sheet["E3"] = os.environ.get("USERNAME")
sheet.cell(row=3,column=5).value='NA'
sheet.cell(row=4,column=5).value=f'{run_time} seconds'
sheet.cell(row=5,column=5).value=stamp
sheet.cell(row=6,column=5).value=status
# ws= xfile.active
# ws.protection.sheet= True
# ws.protection.enable()
# ws.protection.set_password("123")
# xfile.security = openpyxl.workbook.protection.WorkbookProtection(workbookPassword="123",lockStructure=True)
# a= input("done ")
xfile.save(target)
print("Excel Saved")
from xlsx2html import xlsx2html 
# a=path+'Autosapexcel_sheet-'+str(stamp)+'.html'
a=path+'Autosapexcel_sheetlink_'+str(stamp)+'.html'
frm="noreply-autosapautomation@gsk.com"
# To=["erry.8.lavakumar@gsk.com"]
# Cc=["akoju.x.sharanya@gsk.com","erry.8.lavakumar@gsk.com"]
# To = ["ITTCSESAT@gsk.com"]
# Cc = ["vishnuvardhan.x.kalicheti@gsk.com","preetham.x.aranha@gsk.com","anil.x.reddy@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com","shiva.x.prasad@gsk.com","akash.2.verma@gsk.com"]
To = ["vxrepbasis@gsk.com","ITTCSESAT@gsk.com","jay.x.menon@gsk.com","sumeet.x.paprikar@gsk.com"]
Cc = ["bunmi.x.jolaoso@gsk.com","krishnamurthy.x.k@gsk.com","preetham.x.aranha@gsk.com","anil.x.reddy@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com","shiva.x.prasad@gsk.com","akash.2.verma@gsk.com"]
     
# To = ["ittcsesat@gsk.com","anil.x.reddy@gsk.com"] 
# Cc = ["akoju.x.sharanya@gsk.com","erry.8.lavakumar@gsk.com","hareesh.x.s@gsk.com","preetham.x.aranha@gsk.com"]
print("To:",To)
print("CC:",Cc)
# Cc = ["bunmi.x.jolaoso@gsk.com","krishnamurthy.x.k@gsk.com","preetham.x.aranha@gsk.com","anil.x.reddy@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com"]
Bcc = ["akoju.x.sharanya@gsk.com"]#,"nikhil.x.chalikwar@gsk.com","erry.8.lavakumar@gsk.com"
# To="nikhil.x.chalikwar@.com"
sub = f" [{status}] : AutoSAP Wave 4 URL monitoring {stamp} GMT"
xlsx2html(target,a)
print("HTML Saved")
import sender
print(sender.send_mail(send_from=frm, send_to=To, send_cc=Cc, subject=sub, text=None, files=None,html=a))
# print(sender.sender(path=path,To=To,sub=sub,ht=a))
# a=input("enter anything to exit")