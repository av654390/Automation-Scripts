
# ****************************************************************************
#  Program Description : BASIS(rbp&rwp) HOURLY MONITORING                    *
#  Program Name:  RBP\autosaprbp.py                                          *
#          Date:  08/05/2024                                                 *
#       version:  1.0.0                                                      *
#        Author:  Sharanya Akoju                                             *
#  Return Codes:                                                             *
#                 0 - Success                                                *
#                 1 - Error check log file                                   *
# ****************************************************************************
#  AutoSAP Automation                                                        *
#  --------------------                                                      *
#  Sr.         Role           Member          Email                          *
#  ---------   ----------     --------------  -------------------------------*
#  1           Developer      akoju sharanya    akoju.x.sharanya@gsk.com     * 
#  2           Developer      erry lavakumar    erry.8.lavakumar@gsk.com     *   
#                                                                            *
# ****************************************************************************

import time
import datetime
import sys
import os
import pytz
import openpyxl
import shutil
import cc
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html
import rz03
import smlg
import sm58
import smq1
import smq2cf
import st04,st22
import sm14,ldap
import logon
import sender

start = datetime.datetime.now()

# # Set timezone to GMT
tz = pytz.timezone('GMT')

# # Generate timestamp
stamp = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d-%H-%M-%S")

cd=os.getcwd()
path = cd + f'/excel data/AutoSAP_{stamp}/'
os.makedirs(os.path.dirname(path),exist_ok=True)
print(path)

# Redirect stdout to a log file
sys.stdout = open(path + "LOG" + stamp + ".txt", "w")

# Copy the original Excel file
actual = os.getcwd() + '\\report.xlsx'
target = path + 'Autosapexcel_sheet-' + str(stamp) + '.xlsx'
shutil.copyfile(actual, target)

# Load the target Excel file
wb = openpyxl.load_workbook(target)
ws = wb.active
ws = wb['Sheet1']

# Calculate time range
current_time = datetime.datetime.now(tz=tz)
one_hour_ago = current_time - datetime.timedelta(hours=1)

#Read System Details
import json
f=open('sys.txt','r')
syst=f.read()
system_ids=json.loads(syst)
f.close()

red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type='solid')#'0000FF00'
amber_fill = PatternFill(start_color='FFFFBF00', end_color='FFFFBF00', fill_type='solid')


all_status = []
row = 11
c = 3
# Loop through each system ID
sessionlist=[]
nos=1
for i in system_ids:
    try:
        session = logon.Autosap(env=system_ids[i])
        print(session)

        def islogonfine(session):
            if isinstance(session,list):
                if session[0]==-2:
                    sender.error_mail(lid=cc.asid,err=session[1],sid=i)
                    sys.exit()
                elif session[0] == -1:
                    time.sleep(3)
                    session = logon.Autosap(env=system_ids[i])

                if isinstance(session,list):
                    sender.error_mail(lid=cc.asid,err=session[1],sid=i)
                    sys.exit()
            return "Logon is fine"
        print(islogonfine(session))

        sessionlist.append(session)

        val11 = st22.st22(nos=nos, session=session, stamp=stamp, path=path, sid=i)
        nos += 1
        total_dumps = val11.get("Sum") if val11 and val11.get("Sum") is not None else 0
        print("val11 ===", total_dumps)
        ws.cell(row=row, column=c+10, value=total_dumps)
        #if total_dumps is not None:
        if total_dumps < 100:
            ws.cell(row=row, column=c+10).fill = green_fill
            all_status.append("Green")
        elif 100 <= total_dumps <= 200:  
            ws.cell(row=row, column=c+10).fill = amber_fill
            all_status.append("Amber")
        else:  # total_dumps > 200
            ws.cell(row=row, column=c+10).fill = red_fill
            all_status.append("Red")

        val1 = rz03.rz03(nos=nos,session=session,stamp=stamp,path=path,sid=i)
        nos+=1
        active_servers = val1.get("ActiveServers")
        print("val1 ===", active_servers)
        ws.cell(row=row, column=c, value=active_servers)
        if i.lower() == "rbp":
            if active_servers == 12:
                ws.cell(row=row,column=c).fill = green_fill
                all_status.append("Green")
            else:
                ws.cell(row=row,column=c).fill=red_fill
                all_status.append("Red")
        elif i.lower() == "rwp":
            if active_servers == 9:
                ws.cell(row=row, column=c).fill = green_fill
                all_status.append("Green")
            else:
                ws.cell(row=row, column=c).fill = red_fill
                all_status.append("Red")

        val2 = smlg.smlg(nos=nos, session=session, stamp=stamp, path=path,sid=i)
        nos+=1
        max_responsetime = val2.get("Max responsetime")
        print("val2 ===", max_responsetime)
        ws.cell(row=row, column=c+1, value=max_responsetime)
        if max_responsetime is not None:
            if max_responsetime <= 1500:
                ws.cell(row=row, column=c+1).fill = green_fill
                all_status.append("Green")
            elif 1500 < max_responsetime <= 3600:
                ws.cell(row=row, column=c+1).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=c+1).fill = red_fill
                all_status.append("Red")

        val9 =  sm14.sm14(nos,session=session,stamp=stamp,path=path,sid=i)
        nos+=1
        status = val9.get("Status")
        print("val9 ===",status)
        ws.cell(row=row,column=c+8, value=status)
        if status =="Active":
            ws.cell(row=row,column=c+8).fill = green_fill
            all_status.append("Green")
        else:
            ws.cell(row=row,column=c+8).fill=red_fill
            all_status.append("Red")

        val4 = smq1.smq1(nos=nos, session=session, stamp=stamp, path=path, queue="SAP_ALE_YAMMES*",sid=i)
        nos+=1
        Total = val4.get("total")
        print("val4===",Total)
        ws.cell(row=row,column=c+3,value=Total)  
        if Total is not None:
            if Total <= 1000:
                ws.cell(row=row,column=c+3).fill = green_fill
                all_status.append("Green")
            elif 1001 < Total <= 1300:
                ws.cell(row=row, column=c+3).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row,column=c+3).fill = red_fill
                all_status.append("Red") 
        
        val5 = smq1.smq1(nos=nos, session=session, stamp=stamp, path=path, queue="CF*",sid=i)
        nos+=1
        Total = val5.get("total")
        print("val5===",Total)
        ws.cell(row=row,column=c+4,value=Total)
        if Total is not None:
            if Total <= 1000:
                ws.cell(row=row,column=c+4).fill = green_fill
                all_status.append("Green")
            elif 1001 < Total <= 1300:
                ws.cell(row=row, column=c+4).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row,column=c+4).fill = red_fill
                all_status.append("Red")

        val6 = smq2cf.smq2cf(nos=nos, session=session, stamp=stamp, path=path, queue="X*",sid=i)
        nos+=1
        Total = val6.get("total")
        print("val6===",Total)
        ws.cell(row=row,column=c+5,value=Total)
        if Total is not None:
            if  Total <= 100:
                ws.cell(row=row,column=c+5).fill = green_fill
                all_status.append("Green")
            elif 101 < Total <= 150:
                ws.cell(row=row, column=c+5).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row,column=c+5).fill = red_fill
                all_status.append("Red")

        val7 = smq2cf.smq2cf(nos=nos, session=session, stamp=stamp, path=path, queue="CF*",sid=i)
        nos+=1
        Total = val7.get("total")
        print("val7===",Total)
        ws.cell(row=row, column=c+6, value=Total )
        if Total is not None:
            if Total <= 1000:
                ws.cell(row=row, column=c+6).fill = green_fill
                all_status.append("Green")
            elif 1001 < Total <= 1300:
                ws.cell(row=row, column=c+6).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=c+6).fill = red_fill
                all_status.append("Red")
        if i.lower()== 'rbp':  #'rbq' #'nbr' #"rbp"
            queue='RBP_AAE'
        elif i.lower()== 'rwp':  #'rwq' #'nwr' #rwp"
            queue='*'
        val3 = sm58.sm58(nos=nos, session=session, stamp=stamp, path=path, queue=queue,sid=i)
        nos += 1
        entries = val3.get("Entries")
        print("val3 ===", entries)
        ws.cell(row=row, column=c+2, value=entries)
        if entries is not None:
            if entries <= 200:
                ws.cell(row=row, column=c+2).fill = green_fill
                all_status.append("Green")
            elif 200 < entries <= 300:
                ws.cell(row=row, column=c+2).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=c+2).fill = red_fill
                all_status.append("Red")
        if i.lower() == 'rbp': #"rbq" #'nbr'
            val10 =  ldap.ldap(nos,session=session,stamp=stamp,path=path,sid=i)
            nos+=1
            status = val10.get("Status")
            print("val10 ===",status)
            ws.cell(row=row,column=c+9, value=status)
            if status =="Green":
                ws.cell(row=row,column=c+9).fill = green_fill
                all_status.append("Green")
            elif status == "Amber":
                ws.cell(row=row,column=c+9).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row,column=c+9).fill=red_fill
                all_status.append("Red")
        elif i.lower()=='rwp':  #'rwq #nwr
            val8 = st04.st04(nos=nos, session=session, stamp=stamp, path=path,sid=i)
            nos += 1
            status = val8.get("Status")
            #print(f"st04 status for system {i}: {status}")  # Debugging statement
            ws.cell(row=row, column=c+7, value=status)
            if status == "Green":
                ws.cell(row=row, column=c+7).fill = green_fill
                all_status.append("Green")
            else:
                ws.cell(row=row, column=c+7).fill = red_fill
                all_status.append("Red")

        row += 1
        print(logon.logof(session))
    except Exception as e:
        row += 1
        all_status.append("Red")
        error_msg = f"Error in system {i}: {str(e)}"
        print(error_msg)

        sender.error_mail(lid=cc.asid,err=error_msg,sid=i)
        exit()
if 'Red' in all_status:
    overall_status='Red'
elif 'Amber' in all_status:
    overall_status='Amber'
elif len(list(set(all_status)))==1:
    overall_status='Green'
# Iterate through each cell in the worksheet
for j in ws.iter_rows():
    for cell in j:
        # Check if the cell value is 0, then convert it to a string
        if cell.value == 0:
            cell.value = "0"


# Populate time-related data
ws["C3"] = one_hour_ago.strftime("%d.%m.%y") #analysisfromdate
ws["F4"] = current_time.strftime("%H:%M:%S") #analysistotime
ws["C4"] = one_hour_ago.strftime("%H:%M:%S") #analysisfromtime
ws["F3"] = current_time.strftime("%d.%m.%y") #analysistodate
ws["F5"] = cc.asid                           #loginid
ws["C5"] = datetime.datetime.now()-start     #timerequired

from openpyxl.styles import Alignment
# Set the alignment to the left for all these cells
alignment = Alignment(horizontal="left")
for cell in ["C3","F4","C4","F3","F5","C5"]:
    ws[cell].alignment = alignment


# Save the workbook
wb.save(target)
try:
    # Convert the Excel file to HTML
    from xlsx2html import xlsx2html
    a = path + 'Autosapexcel_sheet-' + str(stamp) + '.html'
    xlsx2html(target,a)

    frm = "noreply-autosapautomation@gsk.com"
    # To=["akoju.x.sharanya@gsk.com"]
    # Cc=["akoju.x.sharanya@gsk.com"]
    # Bcc=["akoju.x.sharanya@gsk.com"]
    Cc = ["anil.x.reddy@gsk.com","preetham.x.aranha@gsk.com","ITTCSESAT@gsk.com"]
    To = ["ESAT_BASIS_System_owners@gsk.com"]#["ITTCSESAT@gsk.com"]
    Bcc = ["akoju.x.sharanya@gsk.com","akash.2.verma@gsk.com","shiva.x.prasad@gsk.com"]
    sub = f"[{overall_status}] : Autosap RBP&RWP hourly system monitoring {stamp} GMT"
    file_path = path #r"C:\Users\as776099\OneDrive - GSK\Desktop\codes\RBP\excel data\autosap " + stamp
    print(sender.send_mail(send_from = frm, send_to=To,send_cc=Cc,send_bcc=Bcc,subject=sub, text = None,
     files=file_path, html=a))
except Exception as e:
    print("error while sending mail",e)


