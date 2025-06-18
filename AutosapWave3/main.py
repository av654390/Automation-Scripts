import time
import datetime
import sys
import os
import pytz
import openpyxl
import shutil
import cc
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html
import rz03
import st06
import smlg,rz03
import al08,st22
import db02
import sender
import logon

start = datetime.datetime.now()
tz = pytz.timezone('CET')
stamp = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d-%H-%M-%S")
cd=os.getcwd()
path = cd + f'/excel data/AutoSAP_{stamp}/'
os.makedirs(os.path.dirname(path),exist_ok=True)
print(path)

sys.stdout = open(path + "LOG" + stamp + ".txt", "w")

# Copy the original Excel file
actual = os.getcwd() + '\\template.xlsx'
target = path + 'Autosapexcel_sheet-' + str(stamp) + '.xlsx'
shutil.copyfile(actual, target)

# Load the target Excel file
wb = openpyxl.load_workbook(target)
ws = wb.active
ws = wb['Sheet1']

# Calculate time range
current_time = datetime.datetime.now(tz=tz)
one_hour_ago = current_time - datetime.timedelta(hours=1)

#Read System Details
import json
f=open('sys.txt','r')
syst=f.read()
system_ids=json.loads(syst)
f.close()

red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type='solid')#'0000FF00'
amber_fill = PatternFill(start_color='FFFFBF00', end_color='FFFFBF00', fill_type='solid')

all_status = []
row = 11
c = 1
# Loop through each system ID
sessionlist=[]
nos=1
for i in system_ids:
    try:
        session = logon.Autosap(env=system_ids[i])
        print(session)

        def islogonfine(session):
            if isinstance(session,list):
                if session[0]==-2:
                    sender.error_mail(lid=cc.asid,err=session[1],sid=i)
                    sys.exit()
                elif session[0] == -1:
                    time.sleep(3)
                    session = logon.Autosap(env=system_ids[i])

                if isinstance(session,list):
                    sender.error_mail(lid=cc.asid,err=session[1],sid=i)
                    sys.exit()
            return "Logon is fine"
        print(islogonfine(session))
        sessionlist.append(session)



        val2 = st06.st06(nos=nos, session=session, stamp=stamp, path=path, sid=i)
        nos += 1
        average_cpu = val2.get("AvgCPU")
        free_memory = val2.get("AvgMemory")  
        print(f"CPU Idle Value: {average_cpu}")
        print(f"Free Memory Value: {free_memory}")

        #print("AvgCPU ===", average_cpu)
        ws.cell(row=row, column=c+1, value=average_cpu)
        if average_cpu is not None:
            if average_cpu <= 90:
                ws.cell(row=row, column=c+1).fill = green_fill
                all_status.append("Green")
            elif 90 < average_cpu <= 95:
                ws.cell(row=row, column=c+1).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=c+1).fill = red_fill
                all_status.append("Red")

        #print("AvgMemory ===", free_memory)
        ws.cell(row=row, column=c+2, value=free_memory)
        if free_memory is not None:
            if free_memory <= 90:
                ws.cell(row=row, column=c+2).fill = green_fill
                all_status.append("Green")
            elif 90 < free_memory <= 95:
                ws.cell(row=row, column=c+2).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=c+2).fill = red_fill
                all_status.append("Red")

    
        val4 = smlg.smlg(nos=nos, session=session, stamp=stamp, path=path,sid=i)
        nos+=1
        max_responsetime = val4.get("Max responsetime")
        print(f"MaxResponse_time: {max_responsetime}")
        #print("val4 ===", max_responsetime)
        ws.cell(row=row, column=c+3, value=max_responsetime)
        if max_responsetime is not None:
            if max_responsetime <= 1500:
                ws.cell(row=row, column=c+3).fill = green_fill
                all_status.append("Green")
            elif 1500 < max_responsetime <= 3600:
                ws.cell(row=row, column=c+3).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=c+3).fill = red_fill
                all_status.append("Red")

        val5 = al08.al08(nos=nos, session=session, stamp=stamp, path=path,sid=i)
        nos+=1
        dialogueusers = val5.get("Users")
        print(f"Dialogue_Users: {dialogueusers}")
        #print("val5 ===", dialogueusers)
        ws.cell(row=row, column=c+4, value=dialogueusers)

        val6 = rz03.rz03(nos=nos,session=session,stamp=stamp,path=path,sid=i)
        nos+=1
        active_servers = val6.get("ActiveServers")
        print(f"Application_servers: {active_servers}")
        #print("val6 ===", active_servers)
        ws.cell(row=row, column=c+5, value=active_servers)
        
        val7 = st22.st22(nos=nos, session=session, stamp=stamp, path=path, sid=i)
        nos += 1
        total_dumps = val7.get("Dumps_Count")  
        print(f"Total_Dumps: {total_dumps}")
        if total_dumps is None:
            total_dumps = 0
        else:
            total_dumps = int(total_dumps)
        ws.cell(row=row, column=c+6, value=total_dumps)
        if total_dumps < 100:
            ws.cell(row=row, column=c+6).fill = green_fill
            all_status.append("Green")
        elif 100 <= total_dumps <= 250:
            ws.cell(row=row, column=c+6).fill = amber_fill
            all_status.append("Amber")
        else:
            ws.cell(row=row, column=c+6).fill = red_fill
            all_status.append("Red")
        
        row += 1
        print(logon.logof(session))
    except Exception as e:
        row += 1
        all_status.append("Red")
        error_msg = f"Error in system {i}: {str(e)}"
        print(error_msg)
        sender.error_mail(lid=cc.asid,err=error_msg,sid=i)
        exit()

if 'Red' in all_status:
    overall_status='Red'
elif 'Amber' in all_status:
    overall_status='Amber'
elif len(list(set(all_status)))==1:
    overall_status='Green'
for j in ws.iter_rows():
    for cell in j:
        # Check if the cell value is 0, then convert it to a string
        if cell.value == 0:
            cell.value = "0"

ws["B3"] = one_hour_ago.strftime("%d.%m.%y") #analysisfromdate
ws["F4"] = current_time.strftime("%H:%M:%S") #analysistotime
ws["B4"] = one_hour_ago.strftime("%H:%M:%S") #analysisfromtime
ws["F3"] = current_time.strftime("%d.%m.%y") #analysistodate
ws["F5"] = cc.asid                           #loginid
ws["B5"] = datetime.datetime.now()-start     #timerequired

from openpyxl.styles import Alignment
# Set the alignment to the left for all these cells
alignment = Alignment(horizontal="left")
for cell in ["B3","F4","B4","F3","F5","B5"]:
    ws[cell].alignment = alignment
# Save the workbook
wb.save(target)
try:
    # Convert the Excel file to HTML
    from xlsx2html import xlsx2html
    a = path + 'Autosapexcel_sheet-' + str(stamp) + '.html'
    xlsx2html(target,a)

    frm = "noreply-autosapautomation@gsk.com"
    To = ["erry.8.lavakumar@gsk.com","gayathri.x.chamarthi@gsk.com"]#,"nikhil.x.chalikwar@gsk.com","erry.8.lavakumar@gsk.com"
    Cc = ["akoju.x.sharanya@gsk.com"]
    # To = ["vxrepbasis@gsk.com","ITTCSESAT@gsk.com","jay.x.menon@gsk.com","sumeet.x.paprikar@gsk.com"]
    # Cc = ["bunmi.x.jolaoso@gsk.com","krishnamurthy.x.k@gsk.com","preetham.x.aranha@gsk.com","anil.x.reddy@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com"]
    Bcc = ["akoju.x.sharanya@gsk.com"]#,"nikhil.x.chalikwar@gsk.com","erry.8.lavakumar@gsk.com"
    sub = f"[{overall_status}] : AutoSAP Wave 3 System Monitoring {stamp} CET"
    file_path = path #r"C:\Users\as776099\OneDrive - GSK\Desktop\codes\RBP\excel data\autosap " + stamp
    print(sender.send_mail(send_from = frm, send_to=To,send_cc=Cc,send_bcc=Bcc,subject=sub, text = None, files=file_path, html=a))#send_cc=Cc,
except Exception as e:
    print("error while sending mail",e)

