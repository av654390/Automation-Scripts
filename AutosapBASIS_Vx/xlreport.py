# ****************************************************************************
#  Program Description : Script Generates XL Report and HTML Summary         *
#  Program Name:  AutosapBASIS\xlreport.py                                   *
#          Date:  21/06/2024                                                 *
#       version:  1.0.0                                                      *
#        Author:  Erry Lavakumar                                             *
#  Return Codes:                                                             *
#                 0 - Success                                                *
#                 1 - Error check log file                                   *
# ****************************************************************************
#  AutoSAP Automation                                                        *
#  --------------------                                                      *
#  Sr.         Role           Member          Email                          *
#  ---------   ----------     --------------  -------------------------------*
#  1           Developer      NIKHIL CHALIKWAR  nikhil.x.chalikwar@gsk.com   *
#  2           Developer      erry lavakumar    erry.8.lavakumar@gsk.com     *  
#  2           Developer      akoju sharanya    akoju.x.sharanya@gsk.com     *  
#                                                                            *
# ****************************************************************************

import shutil
import openpyxl
from openpyxl.styles import PatternFill

def genxl(final_values=0,run_time=0,path=0,stamp=0):
    org='report.xlsx'
    tar=path+f'\AutoSAP_XL_{stamp}.xlsx'
    shutil.copyfile(org,tar)
    wb=openpyxl.load_workbook(tar)
    ws=wb['Sheet1']

    redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
    greenFill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')  
    amberFill=PatternFill(start_color='FFA500',end_color='FFA500',fill_type='solid')  
    yellowFill=PatternFill(start_color='FFFF00',end_color='FFFF0000',fill_type='solid')

    import cc
    
    ws.cell(row=3,column=3).value=stamp[:10]
    ws.cell(row=4,column=3).value=f'{stamp[11:]} GMT'
    ws.cell(row=3,column=5).value=cc.elkid
    ws.cell(row=4,column=5).value=f'{run_time} seconds'

    def fill_color(status='',irow=0):  
        if status.lower()=='green':
            ws.cell(row=irow,column=5,value=status).fill=greenFill
        elif status.lower()=='amber':
            ws.cell(row=irow,column=5,value=status).fill=amberFill
        elif status.lower()=='red':
            ws.cell(row=irow,column=5,value=status).fill=redFill
        else:
            ws.cell(row=irow,column=5,value=status).fill=yellowFill

    #####SE38#####
    #CPU Utilization%
    cpu=final_values['se38_dict']['cpu%']['value']
    ws.cell(row=6,column=3).value=str(cpu)
    fill_color(status=final_values['se38_dict']['cpu%']['status'],irow=6)
    #Logins
    logins=final_values['se38_dict']['logins']['value']
    ws.cell(row=20,column=3).value=str(logins)
    fill_color(status=final_values['se38_dict']['logins']['status'],irow=20)
    #Sessions
    sessions=final_values['se38_dict']['sessions']['value']
    ws.cell(row=23,column=3).value=str(sessions)
    fill_color(status=final_values['se38_dict']['sessions']['status'],irow=23)
    #WPS Priv Mode
    wps=final_values['se38_dict']['wps']['value']
    ws.cell(row=29,column=3).value=str(wps)
    fill_color(status=final_values['se38_dict']['wps']['status'],irow=29)

    ######ST06######
    #RAGENT CPU
    ws.cell(row=9,column=3).value=final_values['st06_dict']['rcpu']['value']
    fill_color(status=final_values['st06_dict']['rcpu']['status'],irow=9)
    #Disk Space
    ws.cell(row=37,column=3).value=final_values['st06_dict']['disk']['value']
    fill_color(status=final_values['st06_dict']['disk']['status'],irow=37)

    ######SM37######
    #Active Jobs
    ws.cell(row=15,column=3).value=str(final_values['sm37_dict']['value'])
    fill_color(status=final_values['sm37_dict']['status'],irow=15)

    ######ST07######
    #Active Users
    ws.cell(row=12,column=3).value=final_values['st07_dict']['act']['value']
    fill_color(status=final_values['st07_dict']['act']['status'],irow=12)
    #Work Process Occupied
    ws.cell(row=17,column=3).value=final_values['st07_dict']['wpo']['value']
    fill_color(status=final_values['st07_dict']['wpo']['status'],irow=17)

    ######ST22######
    ws.cell(row=26,column=3).value=final_values['st22_dict']['value']  
    fill_color(status=final_values['st22_dict']['status'],irow=26)
    
    ######DB02######  
    nrow=32
    for i in final_values['db02_dict']:
        ws.cell(row=nrow,column=3).value=final_values['db02_dict'][i]['value']
        fill_color(status=final_values['db02_dict'][i]['status'],irow=nrow)
        nrow+=1
        if nrow==37:
            print(final_values['db02_dict'][i]['value'],nrow)
            break
    
    ######Seal Server(SM59)####
    ws.cell(row=39,column=3).value=final_values['sm59_dict']['value']
    fill_color(status=final_values['sm59_dict']['status'],irow=39)

    # ws.cell(row=40,column=2).value=r'<img src="cid:<C:\Users\el491900\OneDrive - GSK\Desktop\scripts\venv\AutosapBASIS\excel data\AutoSAP_2024-05-21-16-15-45\\db01.png>">'
    wb.save(tar)
    print("XL Generated")
    from xlsx2html import xlsx2html
    ht=path+f'\AutoSAP_html_summary_{stamp}.html'
    xlsx2html(tar,ht)
    print("HTML Generated")
    return ht


def rfcxl(sm59_dict=0,combined_dict=0,path=0,stamp=0,sid=0):
    try:
        org='RFC_Details.xlsx'
        tar=path+f'\\0[RFC]_AutoSAP_XL{stamp}.xlsx'
        shutil.copyfile(org,tar)
        wb=openpyxl.load_workbook(tar)
        ws=wb['Sheet1']
        print(sm59_dict)
        r=2
        for k in sm59_dict:
            ws.cell(row=r,column=3).value=sm59_dict[k]
            r+=1

        ws2=wb['Sheet2']
        x=2
        for k,v in combined_dict.items():
            ws2.cell(row=x,column=1).value=k
            ws2.cell(row=x,column=2).value=str(v[0])
            ws2.cell(row=x,column=3).value=str(v[1])
            x+=1
        wb.save(tar)
        print("RFC XL Generated Successfully!")
    except Exception as e:
        print({"Error while generating RFC Excel":e})


if __name__=='__main__':
    sm59_dict={99: 'Ok', 100: 'Ok', 101: 'Ok', 102: 'Ok', 103: 'Ok', 104: 'Ok', 105: 'Ok', 106: 'Ok', 107: 'Ok', 108: 'Ok', 109: 'Ok', 110: 'Ok', 111: 'Ok', 112: 'Ok', 113: 'Ok', 114: 'Ok', 115: 'Ok', 116: 'Ok', 117: 'Ok', 118: 'Ok', 119: 'Ok', 120: 'Ok', 121: 'Ok', 122: 'Ok', 123: 'Ok', 124: 'Ok', 129: 'Ok'}
    combined_dict= {'rixsap9v_SBT_11': [14, 40.684], 'rixsasap001_SBT_01': [46, 41.301]}
    import os
    path=os.getcwd()
    rfcxl(sm59_dict=sm59_dict,combined_dict=combined_dict,path=path,stamp="Now")
