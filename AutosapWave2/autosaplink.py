import time
import datetime
import pytz
import openpyxl
import shutil
from openpyxl.styles import PatternFill
import os
import sys
import pandas as pd
import linktester
from datetime import timedelta
from xlsx2html import xlsx2html
import sender

start = time.time()

# Timezone setup
# tz=pytz.timezone('Asia/Kolkata')
tz = pytz.timezone('Europe/London')

# Generate timestamp for file naming
stamp = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d-%H-%M-%S")

# Create directory for the Excel output
path = os.getcwd() + f"\\excel data\\autosap {stamp}\\"
os.mkdir(path)

# Redirect stdout to a log file
sys.stdout = open(path + "LOG" + stamp + ".txt", "w")

# Define file paths
original = os.getcwd() + '\\WAVE2.xlsx'
#original = os.getcwd() +'\\og3 - Copy.xlsx'
# original = os.getcwd() +'\javafailed.xlsx'
target = path + 'Autosaplinkexcel_sheet-' + str(stamp) + '.xlsx'
shutil.copyfile(original, target)

# Load the Excel workbook
xfile = openpyxl.load_workbook(target)
sheet = xfile['Sheet1']

# Define cell fill styles for PASS and FAIL
fill_cell_pass = PatternFill(patternType='solid', fgColor='228B22')  # Green
fill_cell_fail = PatternFill(patternType='solid', fgColor='FF4040')  # Red

# Load the data into a pandas DataFrame
df = pd.read_excel(target, header=6, engine='openpyxl')

# Filter out rows where the URL column is empty or NaN
df = df[df['URL'].notna()]

# Initialize failure count
b = 0

# Process each URL in the filtered DataFrame
for i in df.index:
    url = str(df['URL'][i]).strip()  # Convert to string and strip spaces
    result = linktester.linktester(url)  # Check the URL
    
    print(result, i)
    
    # Update Excel with PASS/FAIL status
    if result:
        sheet[f"D{i+8}"] = "PASS"
        sheet[f"D{i+8}"].fill = fill_cell_pass
    else:
        b += 1
        sheet[f"D{i+8}"] = "FAIL"
        sheet[f"D{i+8}"].fill = fill_cell_fail

# Add timestamp, status, and other metadata to the Excel sheet
tm = datetime.datetime.now(tz=tz) - timedelta(hours=1)
sheet["B3"] = tm.strftime("%d.%m.%y")
sheet["B6"] = datetime.datetime.now(tz=tz).strftime("%H:%M:%S")
sheet["B4"] = tm.strftime("%H:%M:%S")
sheet["B5"] = datetime.datetime.now(tz=tz).strftime("%d.%m.%y")
status = "Green" if b == 0 else "Red"
sheet["E6"] = status
end = time.time()
execution_time = round(end - start)
sheet["E4"] = f"{execution_time} sec."
sheet["E5"] = stamp
sheet["E3"] = os.environ.get("USERNAME")

# Protect the Excel sheet with a password
ws = xfile.active
ws.protection.sheet = True
ws.protection.enable()
ws.protection.set_password("123")

# Save the Excel file
xfile.save(target)

# Convert the Excel file to HTML
html_output = path + 'Autosapexcel_sheetlink-' + str(stamp) + '.html'
xlsx2html(target, html_output)

# Email configuration
frm = "noreply-autosapautomation@gsk.com"
to = ["ittcsesat@gsk.com","anil.x.reddy@gsk.com"]
cc = ["akoju.x.sharanya@gsk.com","erry.8.lavakumar@gsk.com","hareesh.x.s@gsk.com","preetham.x.aranha@gsk.com"] #
Bcc = ["akoju.x.sharanya@gsk.com","erry.8.lavakumar@gsk.com"]#
sub = f" [{status}] : AutoSAP URL monitoring {stamp} GMT"

# Send the email
print(sender.send_mail(send_from=frm, send_to=to, send_cc=cc, send_bcc=Bcc, subject=sub, text=None, files=None, html=html_output))

#11General	RPP	https://us6salxrpp204.corpnet2.com:59101/nwa
 
