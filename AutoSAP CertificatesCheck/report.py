import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
import os
import shutil
import datetime
import time
from dateutil.relativedelta import * 
import autosapcheck

def rex(fpath,dict_list,run_time):
    now=datetime.datetime.now()
    end=now+relativedelta(weeks=+1)
    today=now.strftime('%d.%m.%Y')
    original = os.getcwd() +'\\cert_format.xlsx'
    # stamp=time.time()
    target =fpath+f'AutoSAPexcel_sheet-{str(today)}.xlsx'
    shutil.copyfile(original, target)
    #ssprint(target)
    xfile = openpyxl.load_workbook(target)

    # global sheet
    # sheet = xfile.get_sheet_by_name('sheet1')
    sheet = xfile['Sheet1']
    # fill_cell1 = PatternFill(patternType='solid',fgColor='FF4040') #red
    # fill_cell2 = PatternFill(patternType='solid',fgColor='228B22') #green#green
    
    sheet['K3'].value=autosapcheck.id
    sheet['K4'].value=str(run_time)+" Seconds"
    sheet['D3'].value=today
    sheet['D5'].value=end.strftime('%d.%m.%Y')
    sheet['K5'].value=now
    
    def fill(c,i,sid):
        print(f"EXCEL- {sid} ",i)
        sheet.cell(row=c,column=2,value=str(i['overdue']))
        sheet.cell(row=c,column=3,value=str(i['week0']))
        sheet.cell(row=c,column=4,value=str(i['week1']))
        sheet.cell(row=c,column=5,value=str(i['week2']))
        sheet.cell(row=c,column=6,value=str(i['week3']))
        sheet.cell(row=c,column=7,value=str(i['week4']))
        sheet.cell(row=c,column=8,value=str(i['week5']))
        sheet.cell(row=c,column=9,value=str(i['week6']))
        sheet.cell(row=c,column=10,value=i['recent'])
        sheet.cell(row=c,column=11,value=f" {i['certname']}")
        print("In sheet",i['certname'])

    rows=sheet.max_row
    # cols=sheet.max_column
    for ro in range(8,rows+1):
        for sid in dict_list:
            # print(ws.cell(row=ro,column=2).value)
            if sheet.cell(row=ro,column=1).value == sid:
                fill(ro, dict_list[sid], sid)
    xfile.save(target)
    from xlsx2html import xlsx2html

    ht=fpath+'Autosaphtml_sheet-'+today+'.html'
    xlsx2html(target,ht)

    return ht

def report(fpath,key_list,run_time,output_list):
    try:
        
        from wcount import wcount
        dict_list={}   #os.getcwd()+f"//excel_data//AutoSAP_{stamp}//"
        for k in key_list:
            path=fpath+k #os.getcwd()+f"//excel_data//AutoSAP_{stamp}//{system}"
            print("-"*500)
            print("PATHH:",path)
            files=os.listdir(path) ##
            for i in files:
                if i.endswith('.txt'):
                    des=path.replace('/','\\')+"\\"+i
                    print(des)
                    x=wcount(des)
                    dict_list[k]=x
                    break
                        
            else:
                dict_list[k]={'overdue': 0, 'week0': 0, 'week1': 0, 'week2': 0, 'week3': 0, 'week4': 0, 'week5': 0, 'week6': 0, 'recent':"Exception", 'certname':output_list[k]['message']}
        print("DICT LIST:",dict_list)
        ht=rex(fpath,dict_list,run_time)
        return ht
    except Exception as e:
        print({"report error":e})
        return "Error in Report Generation"

if __name__ == "__main__":
    pass