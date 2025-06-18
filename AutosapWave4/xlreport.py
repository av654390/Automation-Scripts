# ****************************************************************************
#  Program Description : Script Generates XL Report and HTML Summary         *
#  Program Name:  AutosapBASIS\xlreport.py                                   *
#          Date:  21/06/2024                                                 *
#       version:  1.0.0                                                      *
#        Author:  Erry Lavakumar                                             *
#  Return Codes:                                                             *
#                 0 - Success                                                *
#                 1 - Error check log file                                   *
# ****************************************************************************
#  AutoSAP Automation                                                        *
#  --------------------                                                      *
#  Sr.         Role           Member          Email                          *
#  ---------   ----------     --------------  -------------------------------*
#  1           Developer      NIKHIL CHALIKWAR  nikhil.x.chalikwar@gsk.com   *
#  2           Developer      erry lavakumar    erry.8.lavakumar@gsk.com     *  
#  2           Developer      akoju sharanya    akoju.x.sharanya@gsk.com     *  
#                                                                            *
# ****************************************************************************

import shutil
import openpyxl
from openpyxl.styles import PatternFill

def rfcxl(sm59_dict=0,path=0,stamp=0,sid=0):
    try:
        org='RFC_Details.xlsx'
        tar=path+f'\\0[RFC]_AutoSAP_XL{stamp}.xlsx'
        shutil.copyfile(org,tar)
        wb=openpyxl.load_workbook(tar)
        ws=wb['Sheet1']
        print(sm59_dict)
        r=2
        for k in sm59_dict:
            ws.cell(row=r,column=3).value=sm59_dict[k]
            r+=1

        wb.save(tar)
        print("RFC XL Generated Successfully!")
    except Exception as e:
        print({"Error while generating RFC Excel":e})


if __name__=='__main__':
    sm59_dict={99: 'Ok', 100: 'Ok', 101: 'Ok', 102: 'Ok', 103: 'Ok', 104: 'Ok', 105: 'Ok', 106: 'Ok', 107: 'Ok', 108: 'Ok', 109: 'Ok', 110: 'Ok', 111: 'Ok', 112: 'Ok', 113: 'Ok', 114: 'Ok', 115: 'Ok', 116: 'Ok', 117: 'Ok', 118: 'Ok', 119: 'Ok', 120: 'Ok', 121: 'Ok', 122: 'Ok', 123: 'Ok', 124: 'Ok', 129: 'Ok'}
    combined_dict= {'rixsap9v_SBT_11': [14, 40.684], 'rixsasap001_SBT_01': [46, 41.301]}
    import os
    path=os.getcwd()
    rfcxl(sm59_dict=sm59_dict,combined_dict=combined_dict,path=path,stamp="Now")
