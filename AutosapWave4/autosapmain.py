import time
import datetime
import sys
import os
import pytz
import openpyxl
import shutil
import cc
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html
import rz03
import st06
import smlg,rz03
import al08,st22
import sm59,db02,lc10
import xlreport
import sender
import logon

start = datetime.datetime.now()
tz = pytz.timezone('CET')
stamp = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d-%H-%M-%S")
cd=os.getcwd()
path = cd + f'/excel data/AutoSAP_{stamp}/'
os.makedirs(os.path.dirname(path),exist_ok=True)
print(path)

sys.stdout = open(path + "LOG" + stamp + ".txt", "w")

# Copy the original Excel file
actual = os.getcwd() + '\\template.xlsx'
target = path + 'Autosapexcel_sheet-' + str(stamp) + '.xlsx'
shutil.copyfile(actual, target)

# Load the target Excel file
wb = openpyxl.load_workbook(target)
ws = wb.active
ws = wb['Sheet1']

# Calculate time range
current_time = datetime.datetime.now(tz=tz)
one_hour_ago = current_time - datetime.timedelta(hours=1)

#Read System Details
import json
f=open('sys.txt','r')
syst=f.read()
system_ids=json.loads(syst)
f.close()

red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type='solid')#'0000FF00'
amber_fill = PatternFill(start_color='FFFFBF00', end_color='FFFFBF00', fill_type='solid')

all_status = []
# row = 8 # for temp 1
row = 11 # for template.xlsx
c = 0
# Loop through each system ID
sessionlist=[]
nos=1
for i in system_ids:
    try:
        session = logon.Autosap(env=system_ids[i])
        print(session)

        def islogonfine(session):
            if isinstance(session,list):
                if session[0]==-2:
                    sender.error_mail(lid=cc.asid,err=session[1],sid=i)
                    sys.exit()
                elif session[0] == -1:
                    time.sleep(3)
                    session = logon.Autosap(env=system_ids[i])

                if isinstance(session,list):
                    sender.error_mail(lid=cc.asid,err=session[1],sid=i)
                    sys.exit()
            return "Logon is fine"
        print(islogonfine(session))
        sessionlist.append(session)

        ### RZ03 ###
        val2 = rz03.rz03(nos=nos,session=session,stamp=stamp,path=path,sid=i)
        nos+=1
        active_servers = val2.get("ActiveServers")
        print(f"Application_servers: {active_servers}")
        #print("val6 ===", active_servers)
        ws.cell(row=row, column=2, value=active_servers)
        
        ### SMLG ###
        val3 = smlg.smlg(nos=nos, session=session, stamp=stamp, path=path,sid=i)
        nos+=1
        max_responsetime = val3.get("Max responsetime")
        print(f"MaxResponse_time: {max_responsetime}")
        #print("val4 ===", max_responsetime)
        ws.cell(row=row, column=3, value=max_responsetime)
        if max_responsetime is not None :
            if i.lower() =='aup':
                if max_responsetime <= 2400:
                    ws.cell(row=row, column=3).fill = green_fill
                    all_status.append("Green")
                elif 2400 < max_responsetime <= 3600:
                    ws.cell(row=row, column=3).fill = amber_fill
                    all_status.append("Amber")
                else:
                    ws.cell(row=row, column=3).fill = red_fill
                    all_status.append("Red")
            else:
                if max_responsetime <= 1500:
                    ws.cell(row=row, column=3).fill = green_fill
                    all_status.append("Green")
                elif 1500 < max_responsetime <= 3600:
                    ws.cell(row=row, column=3).fill = amber_fill
                    all_status.append("Amber")
                else:
                    ws.cell(row=row, column=3).fill = red_fill
                    all_status.append("Red")

        ### ST06 ###
        val4 = st06.st06(nos=nos, session=session, stamp=stamp, path=path, sid=i)
        nos += 1
        average_cpu = val4.get("AvgCPU")
        free_memory = val4.get("AvgMemory")  
        print(f"CPU Idle Value: {average_cpu}")
        print(f"Free Memory Value: {free_memory}")

        #print("AvgCPU ===", average_cpu)
        ws.cell(row=row, column=4, value=average_cpu)
        if average_cpu is not None:
            if average_cpu <= 84:
                ws.cell(row=row, column=4).fill = green_fill
                all_status.append("Green")
            elif 84 < average_cpu < 95:
                ws.cell(row=row, column=4).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=4).fill = red_fill
                all_status.append("Red")

        #print("AvgMemory ===", free_memory)
        ws.cell(row=row, column=5, value=free_memory)
        if free_memory is not None:
            if free_memory <= 84:
                ws.cell(row=row, column=5).fill = green_fill
                all_status.append("Green")
            elif 84 < free_memory < 95:
                ws.cell(row=row, column=5).fill = amber_fill
                all_status.append("Amber")
            else:
                ws.cell(row=row, column=5).fill = red_fill
                all_status.append("Red")

    
        ### AL08 ###        
        val6 = al08.al08(nos=nos, session=session, stamp=stamp, path=path,sid=i)
        nos+=1
        dialogueusers = val6.get("Users")
        print(f"Dialogue_Users: {dialogueusers}")
        #print("val5 ===", dialogueusers)
        ws.cell(row=row, column=6, value=dialogueusers)

        ### ST22 ###
        val7 = st22.st22(nos=nos, session=session, stamp=stamp, path=path, sid=i)
        nos += 1
        total_dumps = val7.get("Dumps_Count")  
        print(f"Total_Dumps: {total_dumps}")
        if total_dumps is None:
            total_dumps = 0
        else:
            total_dumps = int(total_dumps)
        ws.cell(row=row, column=7, value=total_dumps)
        if total_dumps < 100:
            ws.cell(row=row, column=7).fill = green_fill
            all_status.append("Green")
        elif 100 <= total_dumps <= 250:
            ws.cell(row=row, column=7).fill = amber_fill
            all_status.append("Amber")
        else:
            ws.cell(row=row, column=7).fill = red_fill
            all_status.append("Red")
        
        ### SM59 ###
        if i.lower() == 'sbp':
            val8 = sm59.sm59(nos=nos,session=session,path=path,stamp=stamp,sid=i)
            sealcount=0
            for s in val8.values():
                if s.lower()=='ok':
                    pass
                else:
                    sealcount+=1
            
            sealstatus=red_fill if sealcount>0 else green_fill
            if sealcount == 0:
                sealcount = 'Green'
            ws.cell(row=row, column=8, value=sealcount).fill = sealstatus

            xlreport.rfcxl(sm59_dict=val8,path=path,stamp=stamp,sid=i)
                
            ### DB02 ###
            val9 = db02.db02(nos=nos, session=session, stamp=stamp, path=path, sid=i)
            nos += 1
            tablespace= val9.get("Table Space")
            tablespace_status = val9.get("Status")
            print(f"Table Space: {tablespace}")
            #print("cpumemory ===", average_cpu)
            ws.cell(row=row, column=9, value=f"{tablespace}")
            if tablespace is not None:
                if 'Red' in tablespace_status:
                    ws.cell(row=row, column=9).fill = red_fill
                    ws.cell(row=15, column=1, value=f"SBP tablespace PSAPSR3I: team is working to move tables from PSAPR3I to New TableSpace")
                    all_status.append("Red")
                elif 'Amber' in tablespace_status:
                    ws.cell(row=row, column=9).fill = amber_fill
                    ws.cell(row=15, column=1, value=f"SBP tablespace PSAPSR3I: team is working to move tables from PSAPR3I to New TableSpace")
                    all_status.append("Amber")
                elif len(list(set(tablespace_status)))==1 and 'Green' in tablespace_status:
                    ws.cell(row=row, column=9).fill = green_fill
                    all_status.append("Green")

        
        
        ### LCP LC10 from AUP ###
        if i.lower() == 'aup':
            val10 = lc10.lc10(nos=nos, session=session, stamp=stamp, path=path, sid=i)
            operational_status = val10.get("operational_status")
            auto_backup = val10.get("Automatic Log Backup")
            used_data = round(float(val10.get("Used Area Data").replace(',','.')),2)
            used_log = round(float(val10.get("Used Area Log").replace(',','.')),2)

            # lc10_dict = f"[ LC10 Status : {operational_status} ] <br> [ Automatic Log Backup : {auto_backup} ] <br> [ Used Area Data : {used_data} ] <br> [ Used Area Log : {used_log} ]"
            lc10_string = ""
            lc10_status = []
            if operational_status.lower() == 'green':
                lc10_status.append("Green")
                all_status.append("Green")
            elif operational_status.lower() == 'red':
                lc10_status.append("Red")
                all_status.append("Red")
                lc10_string = lc10_string + f"Operational Status : {operational_status} ] <br> "
            if auto_backup.lower().strip() == 'on':
                lc10_status.append("Green")
                all_status.append("Green")
            else:
                lc10_status.append("Red")
                all_status.append("Red")
                lc10_string = lc10_string + f"Automatic Log Backup : {auto_backup} ] <br> "
            if used_data <60:
                lc10_status.append("Green")
                all_status.append("Green")
            elif used_data >=60:
                lc10_status.append("Red")
                all_status.append("Red")
                lc10_string = lc10_string + f"[ Used Area Data : {used_data} ] <br>"
            if used_log < 50:
                lc10_status.append("Green")
                all_status.append("Green")
            elif used_log >=50:
                lc10_status.append("Red")
                all_status.append("Red")
                lc10_string = lc10_string + f"[ Used Log Data : {used_log} ] <br>"

            if lc10_string == "":
                lc10_string = 'Green'
            print("LC10 Overall Status :",lc10_status)
            if 'Red' in lc10_status:
                lc10_status_fill=red_fill
            elif 'Amber' in lc10_status:
                lc10_status_fill=amber_fill
            elif len(list(set(lc10_status)))==1:
                lc10_status_fill=green_fill 
            ws.cell(row=row, column=10,value=lc10_string).fill = lc10_status_fill
        row += 1
        print(logon.logof(session))
    except Exception as e:
        row += 1
        all_status.append("Red")
        error_msg = f"Error in system {i}: {str(e)}"
        print(error_msg)
        sender.error_mail(lid=cc.asid,err=error_msg,sid=i)
        exit()

if 'Red' in all_status:
    overall_status='Red'
elif 'Amber' in all_status:
    overall_status='Amber'
elif len(list(set(all_status)))==1:
    overall_status='Green'
for j in ws.iter_rows():
    for cell in j:
        # Check if the cell value is 0, then convert it to a string
        if cell.value == 0:
            cell.value = "0"

ws["C3"] = one_hour_ago.strftime("%d.%m.%y") #analysisfromdate
ws["I3"] = current_time.strftime("%d.%m.%y") #analysistodate
ws["C4"] = one_hour_ago.strftime("%H:%M:%S") #analysisfromtime
ws["I4"] = current_time.strftime("%H:%M:%S") #analysistotime
ws["I5"] = cc.asid                           #loginid
ws["C5"] = datetime.datetime.now()-start     #timerequired

from openpyxl.styles import Alignment
# Set the alignment to the left for all these cells
alignment = Alignment(wrap_text=True)
# for cell in ["C6","H6","I6","J6","I8","J9"]:
#     ws[cell].alignment = alignment
# Save the workbook
wb.save(target)
try:
    # Convert the Excel file to HTML
    from xlsx2html import xlsx2html
    a = path + 'Autosapexcel_sheet-' + str(stamp) + '.html'
    xlsx2html(target,a)

    frm = "noreply-autosapautomation@gsk.com"
    # To = ["vishnuvardhan.x.kalicheti@gsk.com","manjula.x.sharma@gsk.com","erry.8.lavakumar@gsk.com"]#,"nikhil.x.chalikwar@gsk.com","erry.8.lavakumar@gsk.com"
    # Cc = ["akoju.x.sharanya@gsk.com"]
    To = ["vxrepbasis@gsk.com","ITTCSESAT@gsk.com","jay.x.menon@gsk.com","sumeet.x.paprikar@gsk.com"]
    Cc = ["bunmi.x.jolaoso@gsk.com","krishnamurthy.x.k@gsk.com","preetham.x.aranha@gsk.com","anil.x.reddy@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com","shiva.x.prasad@gsk.com","akash.2.verma@gsk.com"]
    Bcc = ["akoju.x.sharanya@gsk.com"]#,"nikhil.x.chalikwar@gsk.com","erry.8.lavakumar@gsk.com"
    sub = f"[{overall_status}] : AutoSAP Wave 4 System Monitoring {stamp} CET"
    file_path = path #r"C:\Users\as776099\OneDrive - GSK\Desktop\codes\RBP\excel data\autosap " + stamp
    print(sender.send_mail(send_from = frm, send_to=To,send_cc=Cc,send_bcc=Bcc,subject=sub, text = None, files=file_path, html=a))#send_cc=Cc,
except Exception as e:
    print("error while sending mail",e)

