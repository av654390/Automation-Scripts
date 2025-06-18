# ****************************************************************************
#  Program Description : Script Performs HBSM Monitoring                     *
#  Program Name:  AutosapHBSM/autosaphbsm.py                                 *
#          Date:  21/06/2024                                                 *
#       version:  1.0.0                                                      *
#        Author:  Erry Lavakumar                                             *
#  Return Codes:                                                             *
#                 0 - Success                                                *
#                 1 - Error check log file                                   *
# ****************************************************************************
#  AutoSAP Automation                                                        *
#  --------------------                                                      *
#  Sr.         Role           Member          Email                          *
#  ---------   ----------     --------------  -------------------------------*
#  1           Developer      NIKHIL CHALIKWAR  nikhil.x.chalikwar@gsk.com   *
#  2           Developer      erry lavakumar    erry.8.lavakumar@gsk.com     *  
#  2           Developer      akoju sharanya    akoju.x.sharanya@gsk.com     *  
#                                                                            *
# ****************************************************************************

# Importing Required Modules #
import os
import sys
import time
import pytz
from PIL import Image
from fpdf import FPDF
from datetime import datetime,timedelta

start=time.time() ##Execution Start Time##
tz=pytz.timezone('GMT')
now=datetime.now(tz=tz)
stamp=now.strftime("%Y-%m-%d-%H-%M-%S")
sd=now.strftime("%Y-%m-%d")
yes=now-timedelta(days=1)
yest=yes.strftime("%d.%m.%Y")
cd=os.getcwd()
path = cd + f'\excel data\AutoSAP_{stamp}\\'
os.makedirs(os.path.dirname(path),exist_ok=True)

sys.stdout = open(path+"LOG"+stamp+".txt", "w")
print("Files will be stored in the path-",path)

try:
    import logon
    import smq2
    import cc

    #Read System Details
    import json
    f=open('sys.txt','r')
    sys=f.read()
    sid=json.loads(sys)
    f.close()
    # sid={'SBT':"1.2 - R/3 Acceptance [Informal Testing]",'WMT':"A.2 - eWM Test [Informal Testing]"}
    # {"SBV":"1.1 - R/3 Validation [Formal Testing]"}
    nos=0
    output=[]
    for system in sid:
        session=logon.Autosap(sid[system])
        output.append(smq2.smq2(nos=nos,session=session,path=path,stamp=stamp,system=system))
        print("SMQ2 QUEUE STATUS",output)
        nos+=1
        logon.logof(session)
    overall_status='Green' if output[0]['status']=='Green' and output[1]['status']=='Green' else 'Red'
    end=time.time()
    run_time=round(end-start,2)

    try:
        import shutil
        import openpyxl
        from openpyxl.styles import PatternFill
        import xlsx2html
        org='report.xlsx'
        tar=path+f'AutoSAP_Report_{stamp}.xlsx'
        shutil.copy(org,tar)
        wb=openpyxl.load_workbook(tar)
        ws=wb['Sheet1']
        redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
        greenFill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')  
        amberFill=PatternFill(start_color='FFA500',end_color='FFA500',fill_type='solid')

        ws.cell(row=3,column=3).value=stamp[:10]
        ws.cell(row=4,column=3).value=f'{stamp[11:]} GMT'
        ws.cell(row=3,column=6).value=cc.elkid
        ws.cell(row=4,column=6).value=f'{run_time} secs    '
        r=0
        for system in sid:
            ws.cell(row=6+r,column=2).value=system
            ws.cell(row=6+r,column=3).value=str(output[r]['value'])
            ws.cell(row=6+r,column=4).value=str(len(output[r]['backlogs']))
            ws.cell(row=6+r,column=6,value=output[r]['status']).fill=greenFill if output[r]['status']=='Green' else redFill
            r+=1
        a=0
        b=1
        r+=2
        for system in sid:
            backlogs=output[a]['backlogs']
            print(backlogs)
            for backlog in backlogs:
                ws.cell(row=6+r,column=1).value=str(b)
                ws.cell(row=6+r,column=2).value=f" {system}: {backlogs[backlog]}"
                r+=1
                b+=1
            a+=1
        wb.save(tar)
        print("XL Report Generated")
        ht=path+f'\AutoSAP_html_summary_{stamp}.html'
        xlsx2html.xls2html(tar,ht)
        print("HTML SUMMARY Generated")

        #Converting images to PDF
        pdf=FPDF()
        pdf_path=os.path.join(path,'Error_Queue_Screenshots_PDF.pdf')
        image_list=[]
        for images in os.listdir(path):
            if images.endswith('.jpg'):
                image_path=os.path.join(path,images)

                #Add a new page
                pdf.add_page()

                #Set font for the heading
                pdf.set_font("Arial",'B',size=20)

                #Add the image name as heading
                image_name=os.path.splitext(images)[0]
                pdf.cell(0,10, txt=image_name, ln=True, align='C')

                #Add the image below the heading
                pdf.image(image_path, x=10, y=30, w=pdf.w-20)

        pdf.output(pdf_path)

        import sender
        frm =  "noreply-autosapautomation@gsk.com"
        subject=f"[{overall_status}] AutoSAP HBSM Monitoring {stamp} GMT"
        if overall_status=='Green':
            # to = ["erry.8.lavakumar@gsk.com"]
            # cc = ["erry.8.lavakumar@gsk.com"]
            to= ["TECH_HBSM_SUPPORT@gsk.com"]
            cc =  ["preetham.x.aranha@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com"]
        elif overall_status=='Red':
            # to = ["erry.8.lavakumar@gsk.com"]
            # cc = ["erry.8.lavakumar@gsk.com"]
            to= ["nagesh.x.hanumanthappa@gsk.com","TECH_HBSM_SUPPORT@gsk.com","RD.LDUSteam-SAPsupport@gsk.com","madhu.x.eruvaku@gsk.com"]
            cc= ["preetham.x.aranha@gsk.com","hareesh.x.s@gsk.com","erry.8.lavakumar@gsk.com","akoju.x.sharanya@gsk.com"]
        bdy="test mail"
        file_path=path
        html=ht
        sender.send_mail(send_from=frm, send_to=to, send_cc=cc, subject=subject, text=bdy, path=path,html=ht)

    except Exception as e:
        print({"Error While Generating Report/Sending Mail":e})

except Exception as e:
    print({"Error in Global Try":e})
