# ****************************************************************************
#  Program Description : Main file which calls all other functions to execute *
#  Program Name:  AutoSAPABAPdumps/autosapEWAdumps.py                         *
#          Date:  30/04/2025                                                 *
#       Version:  1.0.0                                                       *
#        Author:  Sharanya Akoju                                              *
#  Return Codes:                                                              *
#                 0 - Success                                                 *
#                 1 - Error check log file                                    *
# ****************************************************************************
#  AutoSAP Automation                                                         *
#  --------------------                                                       *
#  Sr.         Role           Member           Email                          *
#  ---------   ----------     --------------   ------------------------------ *
#  1           Developer      Erry Lavakumar    erry.8.lavakumar@gsk.com      *  
#  2           Developer      Akoju Sharanya    akoju.x.sharanya@gsk.com      *  
# ****************************************************************************

import logon
import sys
import st22
import sender
import cc
import datetime
import os
import time 
import openpyxl
import shutil
import pytz
from xlsx2html import xlsx2html

start = datetime.datetime.now()
tz = pytz.timezone('GMT')
stamp = datetime.datetime.now(tz=tz).strftime("%Y-%m-%d-%H-%M-%S")

cd=os.getcwd()
path = cd + f'/excel data/AutoSAP_{stamp}/'
os.makedirs(os.path.dirname(path),exist_ok=True)
sys.stdout = open(path + "LOG" + stamp + ".txt", "w")

actual = os.getcwd() + '\\report(rx).xlsx'
target = path + 'Autosapexcel_sheet-' + str(stamp) + '.xlsx'
shutil.copyfile(actual, target)
# Load the target Excel file
wb = openpyxl.load_workbook(target)
ws = wb.active
ws = wb['Sheet1']

# Calculate time range 
current_time = datetime.datetime.now(tz=tz)
one_hour_ago = current_time - datetime.timedelta(hours=1)

#read system deatails
import json
with open('sys.txt', 'r') as f:
    sid = json.load(f)

row = 6  # Starting row in Excel
SNO = 1  # Column for serial numbers
SID = 2  # Column for SID names
COUNT = 4  # Column for Dumps count
nos = 1  # Serial number counter

for i in sid:
    print(f"Attempting to log on to: {sid[i]}")  # Debugging line
    session = logon.Autosap(sid[i])

#error code#
    # def islogonfine(session):
    #     if isinstance(session,list):
    #         if session[0]==-2:
    #             sender.error_mail(lid=session[1],err=session[2],sid=i)
    #             sys.exit()
    #         elif session[0]==-1:
    #             time.sleep(3)
    #             session=logon.Autosap(sid[i])

    #         if isinstance(session,list):
    #             sender.error_mail(lid=session[1],err=session[2],sid=i)
    #             sys.exit()
    #     return "Logon is Fine"
    # print(islogonfine(session))
# Error code check for SAP logon
    def islogonfine(session):
        if isinstance(session, dict):
            if session["status"] == -2:
                sender.error_mail(lid=session["asid"], err=session["error"], sid=i)
                sys.exit()
            elif session["status"] == -1:
                time.sleep(3)
                session = logon.Autosap(sid[i])

                if isinstance(session, dict) and session["status"] == -2:
                    sender.error_mail(lid=session["asid"], err=session["error"], sid=i)
                    sys.exit()
        return session

    session = islogonfine(session)
    print("Logon is Fine")

    # Call st22 function to download the files
    try:
        result = st22.st22(nos=nos, session=session, system_id=i, stamp=stamp, path=path)
        status = result.get("Dumps_count")
        print("Dumps Count:", status)

        # Write values to Excel
        ws.cell(row=row, column=SNO, value=nos)  # Update serial number
        ws.cell(row=row, column=SID, value=i)  # Update SID
        ws.cell(row=row, column=COUNT, value=status)  # Update Dumps count

        # Log off the system
        logoff_result = logon.logof(session)
        print(logoff_result)
        # Move to the next row and increment serial number
        row += 1
        nos += 1
    except Exception as e:
        print(f"Error processing system {i}: {e}")

ws["F4"] = cc.asid                           #loginid
ws["B4"] = datetime.datetime.now()-start     #timerequired
ws["B3"] = one_hour_ago.strftime("%d.%m.%y") #analysisfromdate
ws["F3"] = current_time.strftime("%d.%m.%y") #analysistodate
from openpyxl.styles import Alignment
# Set the alignment to the left for all these cells
alignment = Alignment(horizontal="left")
for cell in ["F4","B4","F3","B3"]:
    ws[cell].alignment = alignment

# Save the workbook
wb.save(target)
# Convert the Excel file to HTML
from xlsx2html import xlsx2html
a = path + 'Autosapexcel_sheet-' + str(stamp) + '.html'
xlsx2html(target,a)

# Email sending section
try:
    frm = "noreply-autosapautomation@gsk.com"
    To = ["TCS_GSS_ABAP@gsk.com"]
    Cc = ["ramanath.r.patange@gsk.com"]
    Bcc = ["akoju.x.sharanya@gsk.com","erry.8.lavakumar@gsk.com","akash.2.verma@gsk.com","shiva.x.prasad@gsk.com"]
    sub = f"Autosap EWA Dumps report {stamp}"
    html = "bdy.html"
    send_result = sender.send_mail(send_from=frm, send_to=To,send_bcc=Bcc,send_cc=Cc, subject=sub, text=None, files=path, html=a)# 
    print(send_result)
except Exception as e:
    print(f"Error during email sending: {e}")